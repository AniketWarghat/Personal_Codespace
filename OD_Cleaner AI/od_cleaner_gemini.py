"""
OD Cleaner - Optimized Free Workflow
====================================
Flow:
1. Read survey XLSX
2. Extract unique Origin/Destination place names
3. Try direct OpenStreetMap match on raw names
4. Send only unmatched names to Gemini for correction
5. Retry OpenStreetMap on Gemini-corrected names
6. Write Excel output

Output columns:
- Raw Origin / Destination
- OSM Direct Match
- Gemini Corrected
- OSM Final Match

Usage:
    python od_cleaner_free_optimized.py <input.xlsx> [output.xlsx]

Requirements:
    pip install google-genai pandas openpyxl requests
Environment:
    GEMINI_API_KEY=...
"""

import sys
import os
import json
import time
import requests
import pandas as pd
from google import genai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
MODEL_CANDIDATES = ["gemini-2.5-flash"]
BATCH_SIZE = 250
RETRY_LIMIT = 5
RETRY_DELAY = 15

ORIGIN_COLS = [
    "3a4.Trip Origin", "3b4.Trip Origin", "3c3.Trip Origin", "3d3.Trip Origin",
    "3e3.Trip Origin", "3f3.Trip Origin", "3g3.Trip Origin", "3h3.Trip Origin", "3i3.Trip Origin",
]

DEST_COLS = [
    "3a5.Trip Destination", "3b5.Trip Destination", "3c4.Trip Destination", "3d4.Trip Destination",
    "3e4.Trip Destination", "3f4.Trip Destination", "3g4.Trip Destination", "3h4.Trip Destination",
    "3i4.Trip Destination",
]

NULL_VALUES = {"-", "", "nan", "none", "null", "n/a", "na"}
VEHICLE_FILTER = ["Car", "Taxi/Cab"]

OSM_REGION_HINT = "Mumbai, Maharashtra, India"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_EMAIL = "your-email@example.com"
NOMINATIM_DELAY = 1.1


# ── Helpers ───────────────────────────────────────────────────────────────────
def clean_token(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() in NULL_VALUES else s


def extract_places(df: pd.DataFrame) -> list:
    seen = {}
    vehicle_col = "3.Vehicle Type"

    for o_col, d_col in zip(ORIGIN_COLS, DEST_COLS):
        if o_col not in df.columns or d_col not in df.columns:
            continue

        for _, row in df.iterrows():
            if VEHICLE_FILTER and clean_token(row.get(vehicle_col, "")) not in VEHICLE_FILTER:
                continue

            for val in [row.get(o_col, ""), row.get(d_col, "")]:
                v = clean_token(val)
                if v and v.lower() not in seen:
                    seen[v.lower()] = v

    return sorted(seen.values(), key=str.lower)


def extract_json_array(text: str):
    text = text.strip()
    if text.startswith("[") and text.endswith("]"):
        return json.loads(text)

    start = text.find("[")
    end = text.rfind("]")

    if start != -1 and end != -1 and end > start:
        return json.loads(text[start:end + 1])

    raise ValueError("No valid JSON array found in model response")


# ── Gemini ────────────────────────────────────────────────────────────────────
def get_client(api_key: str):
    return genai.Client(api_key=api_key)


def test_model(client, model_name: str) -> bool:
    try:
        client.models.generate_content(
            model=model_name,
            contents="Reply with only: OK"
        )
        print(f"Using model: {model_name}")
        return True
    except Exception as e:
        msg = str(e)
        if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
            print(f"Model available but quota/rate limit hit: {model_name}")
            return True
        print(f"Model not available: {model_name} -> {e}")
        return False


def get_working_model(client):
    print("Checking available Gemini model...")
    for model_name in MODEL_CANDIDATES:
        if test_model(client, model_name):
            return model_name

    print("No compatible Gemini model found for this API key / SDK.")
    sys.exit(1)


def standardise_batch(client, model_name: str, names: list) -> list:
    prompt = f"""You are a Mumbai transport data analyst cleaning an Origin-Destination survey dataset.

For each place name:
- If it is a real, recognisable location in Mumbai or Maharashtra, fix spelling and standardise it
- Use Title Case
- Keep different places distinct
- Do NOT merge separate places
- If it is gibberish, numeric junk, random text, or has no clear real-place match, return exactly: -
- Return exactly one output for each input in the same order

Return ONLY a valid JSON array of strings.
No explanation. No markdown.

Input:
{json.dumps(names, ensure_ascii=False)}
"""

    wait_seconds = RETRY_DELAY

    for attempt in range(1, RETRY_LIMIT + 1):
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=prompt
            )

            raw = (response.text or "").strip()
            raw = raw.replace("```json", "").replace("```", "").strip()
            result = extract_json_array(raw)

            if not isinstance(result, list):
                raise ValueError("Response is not a list")
            if len(result) != len(names):
                raise ValueError(f"Expected {len(names)} items, got {len(result)}")

            return [str(x).strip() if str(x).strip() else "-" for x in result]

        except Exception as e:
            msg = str(e)
            print(f"  Attempt {attempt}/{RETRY_LIMIT} failed: {e}")

            if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
                print(f"  Quota/rate limit hit. Waiting {wait_seconds} seconds...")
                time.sleep(wait_seconds)
                wait_seconds *= 2
            elif attempt < RETRY_LIMIT:
                time.sleep(RETRY_DELAY)

    print("  All retries failed — returning '-' for this batch")
    return ["-" for _ in names]


def ai_standardise_subset(places: list, api_key: str) -> list:
    if not places:
        return []

    client = get_client(api_key)
    model_name = get_working_model(client)

    corrected = []
    total = len(places)

    for i in range(0, total, BATCH_SIZE):
        batch = places[i:i + BATCH_SIZE]
        end = min(i + BATCH_SIZE, total)

        print(f"  Gemini batch {i + 1}-{end} / {total} ...", end=" ", flush=True)
        fixed = standardise_batch(client, model_name, batch)
        corrected.extend(fixed)

        changed = sum(1 for a, b in zip(batch, fixed) if a != b)
        print(f"Done ({changed} corrected)")

    return corrected


# ── OSM ───────────────────────────────────────────────────────────────────────
def choose_best_osm_name(result: dict) -> str:
    address = result.get("address", {})

    priority_keys = [
        "suburb",
        "neighbourhood",
        "quarter",
        "hamlet",
        "village",
        "town",
        "city_district",
        "city",
        "road",
        "commercial",
        "industrial",
        "hospital",
        "building",
    ]

    for key in priority_keys:
        val = address.get(key, "")
        if val and str(val).strip():
            return str(val).strip()

    name = str(result.get("name", "")).strip()
    if name:
        return name

    display_name = str(result.get("display_name", "")).strip()
    if display_name:
        return display_name.split(",")[0].strip()

    return "-"


def validate_with_nominatim(place: str, session: requests.Session) -> str:
    if not place or place.strip() == "-":
        return "-"

    params = {
        "q": f"{place}, {OSM_REGION_HINT}",
        "format": "jsonv2",
        "addressdetails": 1,
        "limit": 1,
        "countrycodes": "in",
        "email": NOMINATIM_EMAIL,
    }

    headers = {
        "User-Agent": f"ODCleaner/1.0 ({NOMINATIM_EMAIL})"
    }

    try:
        response = session.get(
            NOMINATIM_URL,
            params=params,
            headers=headers,
            timeout=20
        )
        response.raise_for_status()
        data = response.json()

        if not data:
            return "-"

        best = data[0]
        return choose_best_osm_name(best)

    except Exception:
        return "-"


def get_osm_matches(places: list, label: str = "OSM matching") -> list:
    matches = []
    cache = {}
    total = len(places)

    with requests.Session() as session:
        print(f"\n{label}...")
        for idx, place in enumerate(places, start=1):
            print(f"  {idx}/{total} ...", end="\r", flush=True)

            key = (place or "").strip().lower()
            if key in cache:
                matches.append(cache[key])
                continue

            match = validate_with_nominatim(place, session)
            cache[key] = match
            matches.append(match)

            time.sleep(NOMINATIM_DELAY)

    print(f"  {total}/{total} ... done")
    return matches


# ── Pipeline ──────────────────────────────────────────────────────────────────
def process_places(places: list, gemini_api_key: str):
    direct_matches = get_osm_matches(places, label="Direct OSM matching on raw names")

    unmatched_raw = []
    unmatched_indices = []

    for idx, (raw, direct) in enumerate(zip(places, direct_matches)):
        if direct == "-":
            unmatched_raw.append(raw)
            unmatched_indices.append(idx)

    print(f"\nDirect OSM matched   : {len(places) - len(unmatched_raw)}")
    print(f"Need Gemini fallback : {len(unmatched_raw)}")

    corrected_all = ["-"] * len(places)
    final_matches = direct_matches[:]

    if unmatched_raw:
        print("\nRunning Gemini only for OSM-unmatched names...")
        corrected_unmatched = ai_standardise_subset(unmatched_raw, gemini_api_key)

        print("\nRunning OSM again on Gemini-corrected names...")
        final_unmatched_matches = get_osm_matches(
            corrected_unmatched,
            label="Final OSM matching on Gemini-corrected names"
        )

        for idx, corr, final in zip(unmatched_indices, corrected_unmatched, final_unmatched_matches):
            corrected_all[idx] = corr
            final_matches[idx] = final

    for idx, direct in enumerate(direct_matches):
        if direct != "-":
            corrected_all[idx] = places[idx]

    return direct_matches, corrected_all, final_matches


# ── Excel styling ─────────────────────────────────────────────────────────────
def style_workbook(wb):
    header_fill = PatternFill("solid", fgColor="1F497D")
    raw_osm_fill = PatternFill("solid", fgColor="FFF2CC")
    corrected_fill = PatternFill("solid", fgColor="E2EFDA")
    final_osm_fill = PatternFill("solid", fgColor="DDEBF7")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    changed_fill = PatternFill("solid", fgColor="C6EFCE")
    invalid_fill = PatternFill("solid", fgColor="FFDACC")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    normal_font = Font(size=10)
    changed_font = Font(size=10, color="375623", bold=True)
    invalid_font = Font(size=10, color="C00000", bold=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            alt = alt_fill if row_idx % 2 == 0 else PatternFill()

            for cell in row:
                col_hdr = str(ws.cell(1, cell.column).value or "")
                val = str(cell.value or "").strip()
                raw_val = str(ws.cell(cell.row, 2).value or "").strip()

                if col_hdr == "OSM Direct Match":
                    if val == "-":
                        cell.fill = invalid_fill
                        cell.font = invalid_font
                    else:
                        cell.fill = raw_osm_fill
                        cell.font = normal_font

                elif col_hdr == "Gemini Corrected":
                    if val == "-":
                        cell.fill = invalid_fill
                        cell.font = invalid_font
                    elif val != raw_val:
                        cell.fill = changed_fill
                        cell.font = changed_font
                    else:
                        cell.fill = corrected_fill
                        cell.font = normal_font

                elif col_hdr == "OSM Final Match":
                    if val == "-":
                        cell.fill = invalid_fill
                        cell.font = invalid_font
                    else:
                        cell.fill = final_osm_fill
                        cell.font = normal_font

                else:
                    cell.fill = alt
                    cell.font = normal_font

                cell.alignment = left
                cell.border = border

        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 4, 14), 45)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 18

    return wb


# ── Output ────────────────────────────────────────────────────────────────────
def build_output(places: list, direct_matches: list, corrected: list, final_matches: list, output_path: str):
    rows_raw = [[i + 1, p] for i, p in enumerate(places)]
    rows_corr = [
        [i + 1, p, d, c, f]
        for i, (p, d, c, f) in enumerate(zip(places, direct_matches, corrected, final_matches))
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(
            rows_raw,
            columns=["S.No", "Raw Origin / Destination"]
        ).set_index("S.No").to_excel(writer, sheet_name="OD_Raw_Unique")

        pd.DataFrame(
            rows_corr,
            columns=[
                "S.No",
                "Raw Origin / Destination",
                "OSM Direct Match",
                "Gemini Corrected",
                "OSM Final Match",
            ]
        ).set_index("S.No").to_excel(writer, sheet_name="OD_Correction")

    wb = load_workbook(output_path)
    style_workbook(wb)
    wb.save(output_path)

    direct_ok = sum(1 for x in direct_matches if x != "-")
    final_ok = sum(1 for x in final_matches if x != "-")
    gemini_used = sum(1 for d in direct_matches if d == "-")

    print(f"\nSaved : {output_path}")
    print(f"Total places         : {len(places)}")
    print(f"Direct OSM matched   : {direct_ok}")
    print(f"Sent to Gemini       : {gemini_used}")
    print(f"Final OSM matched    : {final_ok}")
    print(f"Still unmatched      : {len(places) - final_ok}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python od_cleaner_free_optimized.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = (
        sys.argv[2]
        if len(sys.argv) >= 3
        else os.path.splitext(input_path)[0] + "_OD_Free_Optimized.xlsx"
    )

    if not os.path.exists(input_path):
        print(f"Input file not found: {input_path}")
        sys.exit(1)

    gemini_api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not gemini_api_key:
        print("GEMINI_API_KEY environment variable not set.")
        sys.exit(1)

    print(f"Reading  : {input_path}")
    df = pd.read_excel(input_path)
    print(f"Rows     : {len(df)}")

    places = extract_places(df)
    print(f"Unique place names : {len(places)}")

    if not places:
        print("No valid Origin/Destination place names found.")
        sys.exit(0)

    direct_matches, corrected, final_matches = process_places(places, gemini_api_key)

    print("\nWriting output...")
    build_output(places, direct_matches, corrected, final_matches, output_path)


if __name__ == "__main__":
    main()